# File Name:   FileReader.py
# Author:      Hao Ye
# Version:     V1.0 09/04/2018
# Purpose:     Transform coordinate data from OSGB coordinate to a local coordinate system #
# Usage:       No arguments needed, but GriderBuilder.py must be in same directory         #

import arcpy
import openpyxl
import os.path
from GridBuilder import Coordinate
from GridBuilder import Zones
from GridBuilder import HetBand

# Set up all essential parameters for data processing
# Set up workspace
arcpy.env.workspace = 'C:\Users\haoye\OneDrive - Carnell Support Services Ltd\Desktop\File Space\Project Management\GIS Development\LocalGrid'    # workspace 

# Set up shapefile file path
#fc = arcpy.GetParameterAsText(0)                                                                                                                                                                                                                           # Zone Key 
fc = 'C:\Users\haoye\OneDrive - Carnell Support Services Ltd\Desktop\File Space\Project Management\GIS Development\LocalGrid\LocalPoints.shp'   # Path for script

# Set up configuration file
cg = 'C:\Users\haoye\OneDrive - Carnell Support Services Ltd\Desktop\File Space\Project Management\GIS Development\LocalGrid\Config.xlsx'   

# Set up output shapefile filepath
#of = arcpy.GetParameterAsText(1)
of = "C:\Users\haoye\OneDrive - Carnell Support Services Ltd\Desktop\File Space\Project Management\GIS Development\LocalGrid\PrjedLayer.shp"

# Set up essential variables required for shapefile processing
fields = ['FID','EASTING','NORTHING','COVER_LEVE']                  # Fields extracted from shapefile
checfd = ['LocalEast', 'LocalNorth']                                # Checked existing Fields
instfd = ['FID','LocalEast', 'LocalNorth']                          # Fields needed to be inserted

#zoneKey = arcpy.GetParameterAsText(1) 
zoneKey = 'A17'

# Configure lists to config grids, height band and points
listPoints = []                                                     # Candindate points from shapefile                                   
listLocalGrid = []                                                  # Local grids reference
listHetBand = []                                                    # Height band reference
pntGemtry = []                                                      # List of point geometries

# Config the zones for local grid system
wb = openpyxl.load_workbook(cg)                                     # Read excel file to retrieve grid parameters
sheet = wb['Grid']
for row in range(2, sheet.max_row + 1):   
    zkey = sheet['A'+str(row)].value                                # Zone Value
    staE = sheet['B'+str(row)].value                                # Northing Value
    staN = sheet['C'+str(row)].value                                # Easting Value
    orgE = sheet['D'+str(row)].value                                # Height Value
    orgN = sheet['E'+str(row)].value                                # Height Value
    prjF = sheet['F'+str(row)].value                                # Projection SF Value 
    zoneCell = Zones(zkey, staE, staN, orgE,orgN, prjF)             # Zone Value
    listLocalGrid.append(zoneCell)                                  # Build Grid Value0

# Config the Height bands for local grid system
sheet = wb['Band']
for row in range(2, sheet.max_row + 1):   
    hegt = sheet['A'+str(row)].value                                # Height Value
    elSF = sheet['B'+str(row)].value                                # Northing Value
    heBd = sheet['C'+str(row)].value                                # Elevation Scale Factor
    comt = sheet['D'+str(row)].value                                # Comments
    bandCell = HetBand(hegt, elSF, heBd, comt)                      # Zone Value
    listHetBand.append(bandCell) 

# Extract field data from from shpfile
cursor = arcpy.da.SearchCursor (fc, fields)                         # Read customised field from script start
for row in cursor:
    coords = Coordinate(zoneKey, row[0], row[1], row[2], row[3])
    listPoints.append(coords)
del cursor

# Populate the parameters to the points list
for coord in listPoints:                                            # Check points in the list
    for zone in listLocalGrid:                                      # Check zones in local grid
        if coord.ZoneKey == zone.getKey:                            # Make sure key is same
            coord.setPrjSF = zone.getPrjSF                          # Projection SF
            coord.StartCordE = zone.StartE                          # Zone Start E
            coord.StartCordN = zone.StartN                          # Zone Start N
            coord.OrignCordE = zone.OrignE                          # Zone Origin E
            coord.OrignCordN = zone.OrignN                          # Zone Origin N
 
# Determine average height of points of the site
sum = 0.0
for coord in listPoints:
    sum += coord.Height
aveHeight = sum/len(listPoints)

# Deterimine the height band of points and pupulate band index
for coord in listPoints:
    coord.ElvSF = 0.9999875
    coord.calcuCSF()            #Pupulate CSF values
    coord.calcuESCoord()        #Populate East and North Value                                

# Check if the coordinate fields have been created
fieldList = arcpy.ListFields(fc)
for field in fieldList:
    existingField = False
    if field.name == "LocalEast" or field.name == "LocalNorth":
        existingField = True
    else:
        existingField = False

if existingField == False:
    arcpy.AddField_management(fc,"LocalEast","DOUBLE",15)
    arcpy.AddField_management(fc,"LocalNorth","DOUBLE",15)

for coord in listPoints:
    uc = arcpy.da.UpdateCursor(fc, instfd)
    for row in uc:
        if coord.FID==row[0]:
            print coord.FID, row[0]
            print coord.LclCordE, coord.LclCordN
            newPt = [coord.FID,str(coord.LclCordE),str(coord.LclCordN)]
            uc.updateRow(newPt)
    del uc

# Create a new shapefile with transformed geometries
for coord in listPoints:
    geometry = coord.BuildGeometry(coord)
    pntGemtry.append(geometry)

# Check if the shapefile is existing, if yes update a new shapefile
if os.path.isfile(of):
    os.remove(of)
    arcpy.CopyFeatures_management(pntGemtry,of)
else:
    arcpy.CopyFeatures_management(pntGemtry,of)
